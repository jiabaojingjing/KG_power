import openpyxl
import xlutils
import os
import xlrd                           #导入模块
from xlutils.copy import copy
import re
path = r'C:\Users\86136\Desktop\文档内容整理\文档'
result = []

def wipe_line_break(str):
    return str.replace("\n", "")

def delete_BS(str):
    return str.replace(" ", "")


def get_allfile(cwd):
    global row
    global col
    global folderflag
    global subfolder
    get_dir = os.listdir(cwd)
    for i in get_dir:
        #print(i)
        sub_dir = os.path.join(cwd,i)
        #print(sub_dir)
        if os.path.isdir(sub_dir):
            subfolder=True
            row += 1
            get_allfile(sub_dir)
        else:
            result.append(i)
    return result

def modifyfile(path,filename,str):

    str1=str+"管理规定"
    #str2=str+"细则"
    str2 = "精益化评价细则"
    if str1 in filename and str2 in filename :
        name = filename[filename.find("分册") + 2:filename.find(str2)]
        print(name)
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        # 得到sheet对象
        sheet['A1'] = '设备'
        sheet['A2'] = name
        ## 指定不同的文件名，可以另存为别的文件
        wb.save(path)
    # else:
    #     print(filename)

def modifyfile2(path,filename,str):
    str1 = str + "管理规定"
    #str2 = str + "细则"
    str2 = "精益化评价细则"
    if str1 in filename and str2 in filename:
        name = filename[filename.find("分册") + 2:filename.find(str2)]
        print(name)
        rb = xlrd.open_workbook(path)
        wb = copy(rb)  # 利用xlutils.copy下的copy函数复制
        ws = wb.get_sheet(0)  # 获取表单0

        ws.write(0, 0, '设备')  # 改变（0,0）的值
        ws.write(1, 0, name)  # 增加（1,0）的值
        wb.save(path)  # 保存文件

def getequip():

    filedata = xlrd.open_workbook(r'C:\Users\86136\Desktop\油浸变压器\QGDW1906-2013输变电一次设备缺陷分类标准 - 副本.xlsx')
    filetable = filedata.sheet_by_index(0)

    for row in range(1, filetable.nrows):
        for col in range(filetable.ncols):
            if col<=1:
                value = filetable.cell_value(row, col)
                if type(value) == str:
                    value1 = wipe_line_break(value)
                    value = delete_BS(value1)
                if value == "" or value=="设备类型" or value=="设备种类" or value=="部件" or value=="部件种类" or value=="部位"  or value=="本体":
                    continue
                print(value)


def getentity(filepath):

    filedata = xlrd.open_workbook(filepath)
    filetable = filedata.sheet_by_index(0)
    for row in range(1, filetable.nrows):
        for col in range(1, filetable.ncols):
            value= filetable.cell_value(row, col)
            value = wipe_line_break(value)
            value = delete_BS(value)
            if(0<len(value)<15) and value not in entity_word:
                entity_word.append(value)
            elif len(value)>=15 and value not in entity_content:
                entity_content.append(value)


def getrepeatfile(filepath,filename):

    entityarray=[]
    filedata = xlrd.open_workbook(filepath)
    filetable = filedata.sheet_by_index(0)
    for row in range(1, filetable.nrows):
        for col in range(1, filetable.ncols):
            value = filetable.cell_value(row, col)
            value = wipe_line_break(value)
            value = delete_BS(value)
            if value in repeatedentity and value not in entityarray:
                entityarray.append(value)
    if len(entityarray)>0:
        print("len(entityarray) "+str(len(entityarray)))
        wb = openpyxl.load_workbook(filepath3)
        sheet = wb.active
        row = sheet.max_row
        print(str(row))
        # 得到sheet对象
        sheet.cell(None,row+1,1).value=filename
        for item in entityarray:
            sheet.cell(None,sheet.max_row+1,2).value=item
        wb.save(filepath3)


if __name__ == "__main__":
    # fileset = get_allfile(path)
    # str3="评价"
    # for item in fileset:
    #     filepath = path+"\\"+item
    #
    #     if os.path.splitext(item)[1] == ".xlsx":
    #         modifyfile(filepath, item,  str3)
    #     elif os.path.splitext(item)[1] == ".xls":
    #         modifyfile2(filepath,item,  str3)
    # getequip()

    global entity_word, entity_content,repeatedentity
    entity_word = []
    entity_content =[]
    repeatedentity=[]
    repeat=[]
    fileset = get_allfile(path)
    # for item in fileset:
    #     filepath = path+"\\"+item
    #     getentity(filepath)
    #
    # for item in entity_word:
    #     for content in entity_content:
    #         # print("conent "+content)
    #         if content.find(item)>=0 and item not in repeat:
    #             repeat.append(item)
    # for i in repeat:
    #     print(i)

    filepath2=r'C:\Users\86136\Desktop\文档内容整理\设备及部件名词.xlsx'
    filepath3=r'C:\Users\86136\Desktop\文档内容整理\模糊的实体名词.xlsx'

    filedata = xlrd.open_workbook(filepath2)
    filetable = filedata.sheet_by_index(2)
    for row in range(0, filetable.nrows):
        for col in range(0, filetable.ncols):
            value =filetable.cell_value(row, col)
            if value not in repeatedentity:
                repeatedentity.append(value)
    print(len(repeatedentity))
    for item in fileset:
        filepath = path+"\\"+item
        getrepeatfile(filepath,item)