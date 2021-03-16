#!/usr/bin/env python3
# coding: utf-8
# File: globalFunction.py

from py2neo import Graph, Node, Relationship, NodeMatcher
import xlrd
import os
import math
import re
import openpyxl
filearray =[]
entityarray = []
partdic={}
equiparray=[]
dicarray=[]
def wipe_line_break(str):
        return str.replace("\n", "").replace(" ", "")

def connectNeo4j():
    file_graph = Graph(
        "http://localhost:7474",
        username="neo4j",
        password="123"
    )
    return file_graph

def getpartdic(filepath):
    filedata = xlrd.open_workbook(filepath)
    filetable = filedata.sheet_by_index(6)

    global equip
    for row in range(0, filetable.nrows):
        for col in range(0, filetable.ncols):
            value = filetable.cell_value(row, col)
            if type(value) == str:
                value = wipe_line_break(value)
            if value == "":
                continue
            if col==0:
                equip=value
                partdic[equip]=[]
            elif col==1:
                part=value
                if part not in  partdic[equip]:
                    partdic[equip].append(part)
    return partdic
def getdicbyType(filepath,sheetname):
    dicarray.clear()
    filedata = xlrd.open_workbook(filepath)
    filetable = filedata.sheet_by_name(sheetname)

    for row in range(0, filetable.nrows):
        for col in range(0, filetable.ncols):
            value = filetable.cell_value(row, col)
            if type(value) == str:
                value = wipe_line_break(value)
            if value == "":
                continue
            if col == 0:
                if value not in dicarray:
                    dicarray.append(value)
    return dicarray

#获取检修策略与评价的相关知识点
def  getfileKnowlodgePoint(filepath):
    pointarray=[]
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.get_sheet_by_name('Sheet1')
    for row in range(2, sheet.max_row):
        for col in range(1, sheet.max_column):
            value=sheet.cell(None,row,col).value
            # if type(value) == str:
                # value = wipe_line_break(value)
            if value == "" or  value is None :
                continue
            attribute = sheet.cell(None,1,col).value
            # attribute = wipe_line_break(attribute)
            if attribute == "关系":
                if value=="关键工艺质量控制":#策略换成“实际状态”
                   point= sheet.cell(None,row,col-1).value
                   if point == "" or point is None:
                       continue
                   else:
                       if point not in pointarray:
                           print(point)
                           pointarray.append(point)
    wb.save(filepath)

def extractRelation(filepath):
    print(filepath)
    chartrow = 300
    chartcol = 300
    global relarray,entityarray
    entityarray.clear()
    filedata = xlrd.open_workbook(filepath)
    filetable = filedata.sheet_by_index(0)
    for row in range(1, filetable.nrows):
        for col in range(0, filetable.ncols):
            # print(str(row), "  ", str(col))
            value = filetable.cell_value(row, col)
            if type(value) == str:
                value = wipe_line_break(value)
            if value == "":
                continue
            attribute = filetable.cell_value(0, col)
            attribute = wipe_line_break(attribute)
            if attribute == "实体":
                entityMatch = re.search("附录", value)
                if entityMatch:
                    chartrow=row
                    chartcol=col
                if col>=chartcol and row>=chartrow and len(value)>0 and len(value)<30 and "。" not in value and "#" not in value:
                    print(value)
               # if value not in entityarray and len(value)<15:
               #     entityarray.append(value)
            # if attribute == "关系":
            #     if value not in relarray:
            #         relarray.append(value)
            #         print(value)
            #         # print(str(row),str(col))
            # if attribute == "实体":#提取regulation类
            #     if value not in  partdic["油浸式变压器（电抗器）"] :
            #
            #         print(value)
def get_allfile(cwd):
    global row
    global col
    global folderflag
    global subfolder
    get_dir = os.listdir(cwd)
    for i in get_dir:
        # print(i)
        sub_dir = os.path.join(cwd,i)
        #print(sub_dir)
        if os.path.isdir(sub_dir):
            subfolder=True
            row += 1
            get_allfile(sub_dir)
        else:

            filearray.append(cwd+i)
    return filearray

def get_equip(path):
    global  entityarray
    filedata = xlrd.open_workbook(path)
    filetable = filedata.sheet_by_index(0)
    for row in range(0, filetable.nrows):
        for col in range(0, filetable.ncols):
            value = filetable.cell_value(row, col)
            if value  not in equiparray:
                equiparray.append(value)

def extractEntity():
    global filearray
    path = r".\文档\\"
    equippath=r".\文档\equip.xlsx"

    get_allfile(path+r"变电检修管理规定细则\\")
    get_allfile(path+r"变电运维管理规定细则\\")
    get_allfile(path+r"变电评价管理规定细则\\")
    get_allfile(path+r"变电验收管理规定细则\\")

    get_equip(equippath)

    # for equip in equiparray:
    equip="站用交流电源系统"
    for file in filearray:
        equipfileMatch = re.search(equip, file)
        if equipfileMatch:
            print("equip: "+equip +" "+file)
            # filepath=path+file
            extractRelation(file)

#
if __name__ == '__main__':
    # path=r"D:\知识图谱\油浸式变压器\equip.xlsx"
    # transformerpath = r"D:\知识图谱\油浸式变压器"
    # transformerarray = []
    # # transformerarray.append(transformerpath + "\国家电网公司变电评价管理规定（试行） 第1分册 油浸式变压器（电抗器）精益化评价细则.xlsx")
    # # transformerarray.append(transformerpath + "\国家电网公司变电检修管理规定（试行） 第1分册 油浸式变压器（电抗器）检修细则.xlsx")
    # # transformerarray.append(transformerpath + "\国家电网公司变电验收管理规定（试行） 第1分册  油浸式变压器（电抗器）验收细则.xlsx")
    # # transformerarray.append(transformerpath + "\国家电网公司变电运维管理规定（试行） 第1分册  油浸式变压器（电抗器）运维细则.xlsx")
    # transformerarray.append(r"D:\知识图谱\206\文档\国家电网公司变电检修管理规定（试行） 第2分册 断路器检修细则.xlsx")
    # getpartdic(path)
    # for pathitem in transformerarray:
    #     getfileKnowlodgePoint(pathitem)

    # extractEntity()
    path = r".\文档"
    servicefilearray = []
    evaluatefilearray = []
    acceptancefilearray = []
    operationsfilearray = []
    detectionfilearray = []

    # detectionfilearray = get_allfile(path + r"\变电检测管理规定细则\\")
    # servicefilearray = get_allfile(path + r"\变电检修管理规定细则\\")
    # operationsfilearray = get_allfile(path + r"\变电运维管理规定细则\\")
    # acceptancefilearray = get_allfile(path + r"\变电验收管理规定细则\\")
    evaluatefilearray = get_allfile(path + r"\规范文档\\")

    for file in detectionfilearray:
        extractRelation(file)
    for file in servicefilearray:
        extractRelation(file)
    for file in operationsfilearray:
        extractRelation(file)
    for file in acceptancefilearray:
        extractRelation(file)
    for file in evaluatefilearray:
        extractRelation(file)
