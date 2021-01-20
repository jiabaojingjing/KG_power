from py2neo import Graph, Node, Relationship, NodeMatcher
import xlrd
import os
from copy import deepcopy
import re
import openpyxl
import xlsxwriter    #调用模块
from xlutils.copy import copy
# import jpype
# from jpype import *
from pyhanlp import *
keyworddic={}
kwawarry=[]
fuzzykwdic={}
fuzzykwawarry=[]
segmentlist=[]
equiplist=[]
relarray=[]
patterndic={}
file_graph = Graph(
    "http://localhost:7474",
    username="neo4j",
    password="123"
)
symbollist=["(",")","（","）"]
lists = [[] for i in range(29)]  #
pattern = [r"验收分类有哪些",r"检修分类有哪些",r"验收要求有哪些",r"异常处置有哪些",r"参加人员有哪些",r"专业巡视要点有哪些",r"关键工艺质量控制有哪些",r"例行检查关键工艺质量控制有哪些",r"更换检修关键工艺质量控制有哪些"
            ,r"更换关键工艺质量控制",r"通用部分检修关键工艺质量控制有哪些",r"安全注意事项有哪些",r"评判项目有哪些",r"评判小项有哪些",r"检查方式有哪些",r"扣分原则有哪些",r"现场检查有哪些",r"运行规定有哪些",r"运行温度要求有哪些",
           r"运行电压要求有哪些",r"并列运行的基本条件是什么",r"紧急申请停运规定是什么",r"巡视内容有哪些",r"操作内容有哪些",r"操作要求是什么",r"维护操作有哪些",r"典型故障有哪些",r"处理原则有哪些",r"[\u4e00-\u9fa5]现象有哪些"]

def wipe_line_break(str):
    return str.replace("\n", "").replace(" ", "")

def getRelation(filepath):
    global equiplist,relarray
    filedata = xlrd.open_workbook(filepath)
    filetable = filedata.sheet_by_index(0)
    for row in range(0, filetable.nrows):
        for col in range(0, filetable.ncols):
            value = filetable.cell_value(row, col)
            if type(value) == str:
                value = wipe_line_break(value)
            if value == "":
                continue
            relarray.append(value)
def getpatterndic():
    for item in pattern:
        patterndic[item]=""

def setpatterndic():
    listsid=0
    for item in pattern:
        if listsid<29:
            patterndic[item]=deepcopy(lists[listsid])
            listsid+=1

    path=r"C:\Users\86136\Desktop\文档内容整理\文档\油浸式变压器\relation2.xlsx"

    wb = openpyxl.load_workbook(path)
    sheet = wb.active

    rowid=1
    for item in patterndic:
        list=patterndic[item]
        sheet.cell(None, rowid, 1).value = item
        # print("问题 ："+item)
        for i in list:
            print(i)
            sheet.cell(None, rowid, 2).value=i
            rowid += 1
    wb.save(path)
def getEntity():
    entitylist=[]
    path = r"C:\Users\86136\Desktop\文档内容整理\文档\油浸式变压器\relation2.xlsx"
    filedata = xlrd.open_workbook(path)
    filetable = filedata.sheet_by_index(0)
    for row in range(0, filetable.nrows):
        value=filetable.cell_value(row, 1)
        if type(value) == str:
            value = wipe_line_break(value)
        if value == "":
            continue
        if value not in entitylist:
            entitylist.append(value)
            print(value)

def getRelationPreKW(filepath):
    global lists,patterndic
    id=0
    filedata = xlrd.open_workbook(filepath)
    filetable = filedata.sheet_by_index(0)
    for item in pattern:
        for row in range(0, filetable.nrows):
            for col in range(0, filetable.ncols):
                value = filetable.cell_value(row, col)
                if type(value) == str:
                    value = wipe_line_break(value)
                if value == "":
                    continue
                attribute = filetable.cell_value(0, col)
                attribute = wipe_line_break(attribute)
                if attribute == "关系":
                    # print(value,"    ",item)
                    match=re.search(value,item)
                    if match != None:
                        # print("match:"+str(match))
                        prevalue = filetable.cell_value(row, col-1)
                        if prevalue=="":
                            prerow=row-1
                            while prerow>0 and prevalue=="":
                                prevalue=filetable.cell_value(prerow, col-1)
                                prerow-=1
                        prevalue = wipe_line_break(prevalue)
                        # if value=="异常处置":
                        #     print("异常处置$$%%   "+prevalue)
                        if prevalue not in lists[id]:
                            lists[id].append(prevalue)

        id+=1
def getEquiplist():
    path = r"C:\Users\86136\Desktop\油浸式变压器\equip.xlsx"
    filedata = xlrd.open_workbook(path)
    filetable = filedata.sheet_by_index(0)

    for row in range(0, filetable.nrows):
        value = filetable.cell_value(row, 0)
        if type(value) == str:
            value = wipe_line_break(value)
        if value == "":
            continue
        if value not in equiplist:
            equiplist.append(value)


# 问题中包含文档中所列关键词
def questionAnswering(question):
    pos = -1
    reldic={}
    tem = []
    entityarray=[]
    # searchedRellist=[]
    entitylist = []
    global segmentlistlength,segmentlist,relarray
    relid=0
    cut_statement = HanLP.segment(question)
    for i in range(len(pattern)):
        # print(i)
        match=re.search(pattern[i],question)
        if(match):
            pos=match.span()[0]
            q_type = i
            break
        if (pos != -1):
            break
    #验收分类
    if (q_type == 0 ):
        index=0
        for term in cut_statement:
            if index>pos:
                break
                index += len(term.word)
            if term.word  in symbollist :
                continue
            if str(term.nature) != "nz":
                continue
            for equip in equiplist:
                term_match=re.search(term.word,equip)
                if term_match:
                    print(term.word)
                    print(equip)
                    print(term_match)
                    term_pos = term_match.span()[0]
                    if term_pos>=0:
                        searchinfo = "MATCH p =(:powerentity{name:" + "'" + equip + "'" + "})" + "-[r:"+r"`验收分类`"+"]->() return p"
                        print(searchinfo)
                        print(file_graph.run(searchinfo))
                        if file_graph.run(searchinfo)!="":
                            break
    # 异常处理
    if (q_type == 2):









    # else:

        # for term in HanLP.segment(question):
        #     # print(term)
        #     if str(term.nature)=="n" or str(term.nature)=="nz":
        #         segmentlist.append(term.word)
        #
        # segmentlistlength= len(segmentlist)
        # while segmentlistlength>0:
        #     # print(str(segmentlistlength))
        #     # print(fuzzysearch(segmentlistlength,filename))
        #     if len(fuzzysearch(segmentlistlength, filename)) >0:
        #         break
        #     segmentlistlength = segmentlistlength - 1



if __name__ == "__main__":
    # getRelation(r"C:\Users\86136\Desktop\文档内容整理\文档\油浸式变压器\relation.xlsx")
    # getRelationPreKW(r"C:\Users\86136\Desktop\文档内容整理\文档\油浸式变压器\国家电网公司变电验收管理规定（试行） 第1分册  油浸式变压器（电抗器）验收细则.xlsx")
    # getRelationPreKW(r"C:\Users\86136\Desktop\文档内容整理\文档\油浸式变压器\国家电网公司变电检修管理规定（试行） 第1分册 油浸式变压器（电抗器）检修细则.xlsx")
    # getRelationPreKW(r"C:\Users\86136\Desktop\文档内容整理\文档\油浸式变压器\国家电网公司变电评价管理规定（试行） 第1分册 油浸式变压器（电抗器）精益化评价细则.xlsx")
    # getRelationPreKW(r"C:\Users\86136\Desktop\文档内容整理\文档\油浸式变压器\国家电网公司变电运维管理规定（试行） 第1分册  油浸式变压器（电抗器）运维细则.xlsx")
    # setpatterndic()
    getEquiplist()
    questionAnswering("油浸式变压器（电抗器）的验收分类有哪些")