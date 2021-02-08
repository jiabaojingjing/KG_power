from py2neo import Graph, Node, Relationship, NodeMatcher
import xlrd
import os
import math
import re
import openpyxl
result = []
row=0
col=0
reldic={}
entitydic={}
relarray=[]
entityarray=[]
chartarray=[]
filearray=[]
regulationarray=[]
file_graph = Graph(
    "http://localhost:7474",
    username="neo4j",
    password="123"
)
path=r"C:\Users\86136\Desktop\文档内容整理\文档"
def wipe_line_break(str):
    return str.replace("\n", "").replace(" ", "")

def extractRelation(filepath):
    # print(filepath)
    global relarray,entityarray
    filedata = xlrd.open_workbook(filepath)
    filetable = filedata.sheet_by_index(0)
    for row in range(1, filetable.nrows):
        for col in range(0, filetable.ncols):
            # print(str(row), "  ", str(col))
            value = filetable.cell_value(row, col)
            if type(value) == str:
                value = wipe_line_break(value)
            if value == "":
                continue
            attribute = filetable.cell_value(0, col)
            attribute = wipe_line_break(attribute)
            # if attribute == "实体":
            #    if value not in entityarray and len(value)<20:
            #        entityarray.append(value)
            #        print(value)
            # if attribute == "关系":
            #     if value not in relarray:
            #         relarray.append(value)
            #         print(value)
            #         # print(str(row),str(col))
            if attribute == "实体":#提取regulation类
                if "附录" in value and len(value)<=20:
                    return
                if value not in entityarray and len(value) < 20:
                    entityarray.append(value)
                    print(value)


def getEntityType():
    global chartarray,filearray,regulationarray
    filedata = xlrd.open_workbook(r"D:\知识图谱\油浸式变压器\字典.xlsx")
    for sheetid in range(0, 3):
        filetable = filedata.sheet_by_index(sheetid)
        for row in range(0, filetable.nrows):
            for col in range(0, filetable.ncols):
                # print(str(row), "  ", str(col))
                value = filetable.cell_value(row, col)
                if type(value) == str:
                    value = wipe_line_break(value)
                if value == "":
                    continue
                if sheetid==0:
                    if value not in chartarray:
                        chartarray.append(value)
                elif sheetid == 1:
                    if value not in regulationarray:
                        regulationarray.append(value)
                elif sheetid == 2:
                    if value not in filearray:
                        filearray.append(value)



def saveFilecontentToNeo4j(filepath):

    global document,documentnode,entitydic,reldic,relnum,entitynum,valueType

    reldic.clear()
    entitydic.clear()
    filedata = xlrd.open_workbook(filepath)
    filetable = filedata.sheet_by_index(0)
    matcher = NodeMatcher(file_graph)
    relnum = 0
    entitynum = 0
    for row in range(1,filetable.nrows):
        for col in range(0,filetable.ncols):
            # print(str(row),"  ",str(col))
            value = filetable.cell_value(row, col)
            if type(value) == str:
                value = wipe_line_break(value)
            if value == "":
                continue

            attribute =  filetable.cell_value(0, col)
            attribute = wipe_line_break(attribute)
            if attribute=="实体":
                name = ""
                desc = ""
                #判断实体类型
                valueType = ""
                if value in regulationarray:
                    valueType = "regulation"
                elif value in filearray:
                    valueType = "file"
                elif value in chartarray:
                    valueType = "chart"
                # if len(valueType) == 0:
                #     for item in chartarray:
                #         chart_match = re.search(item, value)
                #         if chart_match:
                #             valueType = "chart"
                if len(valueType)==0:
                    valueType = "detail"
                relid = int(col/2)
                entityid= math.ceil(col/2)
                if len(reldic)>0 and len(entitydic)>0:
                    if col>0:
                        fathernode=entitydic[entityid-1]
                        relation=reldic[relid-1]
                        if relation!="" and fathernode!="":
                           if valueType=="detail":

                               if len(value.split("#"))>=2:
                                   name=value.split("#")[0]
                                   desc=value.split("#")[1]
                           else:
                               name=value
                        newnode = Node(valueType, name=name, desc=desc)
                        file_graph.create(Relationship(fathernode, relation, newnode))
                        entitydic[entityid]=newnode
                    elif col==0:
                        reldic.clear()
                        entitydic.clear()
                        newnode = matcher.match(valueType, name=value).first()
                        if (newnode is None):
                            newnode = Node(valueType, name=value)
                            file_graph.create(newnode)
                        entitydic[entityid] = newnode
                elif len(reldic)==0 and len(entitydic)==0:
                    newnode = matcher.match(valueType, name=value).first()
                    if (newnode is None):
                        newnode = Node(valueType, name=value)
                        file_graph.create(newnode)
                    entitydic[entityid]=newnode
            elif attribute=="关系":
                reldic[int(col/2)]=value

def modefyfilename():
    data = xlrd.open_workbook(r"D:\知识图谱\油浸变压器\powerspec.xls")
    table = data.sheet_by_index(0)
    for row in range(table.nrows):

        for col in range(table.ncols):
            value = table.cell_value(row, col)
            if col==3 and row>0:
                value=wipe_line_break(value)
                print(os.path.splitext(value)[0])

def modifyfilecontent(filepath):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active
    for row in range(2, sheet.max_row):
        for col in range(1, sheet.max_column):
            value=sheet.cell(None,row,col).value
            if type(value) == str:
                value = wipe_line_break(value)
            if value == "" or  value is None :
                continue
            attribute = sheet.cell(None,1,col).value
            attribute = wipe_line_break(attribute)
            if attribute == "实体":
                print(value)
                if len(value.split("#"))>= 2 and col >= 2:
                    rel = sheet.cell(None,row,col-1).value
                    if rel == "包含":
                        sheet.cell(None, row, col-1).value="相关条款"
    wb.save(filepath)

    # filedata = xlrd.open_workbook(filepath)
    # filetable = filedata.sheet_by_index(0)
    # for row in range(1,filetable.nrows):
    #     for col in range(0,filetable.ncols):
    #         value = filetable.cell_value(row, col)
    #         if type(value) == str:
    #             value = wipe_line_break(value)
    #         if value == "":
    #             continue
    #
    #         attribute = filetable.cell_value(0, col)
    #         attribute = wipe_line_break(attribute)
    #         if attribute == "实体":
    #             if len(value.split("#")) >= 2 and col>=2:
    #                 rel= filetable.cell_value(row, col-1)
    #                 if rel=="含义":



if __name__ == "__main__":

    rulepath=r"C:\Users\86136\Desktop\六项管理规定"
    # saveFilecontentToNeo4j(r"C:\Users\86136\Desktop\文档内容整理\文档\油浸式变压器\国家电网公司变电验收管理规定（试行） 第1分册  油浸式变压器（电抗器）验收细则.xlsx", "")
    # saveFilecontentToNeo4j(r"C:\Users\86136\Desktop\文档内容整理\文档\油浸式变压器\国家电网公司变电检修管理规定（试行） 第1分册 油浸式变压器（电抗器）检修细则.xlsx", "")
    # saveFilecontentToNeo4j(r"C:\Users\86136\Desktop\文档内容整理\文档\油浸式变压器\国家电网公司变电评价管理规定（试行） 第1分册 油浸式变压器（电抗器）精益化评价细则.xlsx", "")
    # saveFilecontentToNeo4j(r"C:\Users\86136\Desktop\文档内容整理\文档\油浸式变压器\国家电网公司变电运维管理规定（试行） 第1分册  油浸式变压器（电抗器）运维细则.xlsx", "")
    # saveFilecontentToNeo4j(r"C:\Users\86136\Desktop\评价表.xlsx")


    # saveFilecontentToNeo4j(r"C:\Users\86136\Desktop\油浸式变压器\国家电网公司变电检修管理规定（试行） 第1分册 油浸式变压器（电抗器）检修细则.xlsx")
    # saveFilecontentToNeo4j(r"C:\Users\86136\Desktop\油浸式变压器\国家电网公司变电评价管理规定（试行） 第1分册 油浸式变压器（电抗器）精益化评价细则.xlsx" )
    # saveFilecontentToNeo4j(r"C:\Users\86136\Desktop\油浸式变压器\国家电网公司变电运维管理规定（试行） 第1分册  油浸式变压器（电抗器）运维细则.xlsx" )
    # extractRelation(r"C:\Users\86136\Desktop\油浸式变压器\国家电网公司变电检修管理规定（试行） 第1分册 油浸式变压器（电抗器）检修细则.xlsx")
    # getEntityType()
    # saveFilecontentToNeo4j(r"C:\Users\86136\Desktop\国家电网公司变电评价管理规定（试行）.xlsx")
    # extractRelation(rulepath+r"\国家电网公司变电检测管理规定（试行）.xlsx")
    # extractRelation(rulepath+r"\国家电网公司变电检修管理规定（试行）.xlsx")
    # extractRelation(rulepath+r"\国家电网公司变电评价管理规定（试行）.xlsx")
    # extractRelation(rulepath+r"\国家电网公司变电验收管理规定（试行）.xlsx")
    # extractRelation(rulepath+r"\国家电网公司变电运维管理规定（试行）.xlsx")
    # extractRelation(rulepath+r"\国家电网公司变电运维检修管理办法（试行）.xlsx")

    # getEntityType()
    # saveFilecontentToNeo4j(rulepath+r"\国家电网公司变电检测管理规定（试行）.xlsx")
    # saveFilecontentToNeo4j(rulepath+r"\国家电网公司变电检修管理规定（试行）.xlsx")
    # saveFilecontentToNeo4j(rulepath+r"\国家电网公司变电评价管理规定（试行）.xlsx")
    # saveFilecontentToNeo4j(rulepath+r"\国家电网公司变电验收管理规定（试行）.xlsx")
    # saveFilecontentToNeo4j(rulepath+r"\国家电网公司变电运维管理规定（试行）.xlsx")
    # saveFilecontentToNeo4j(rulepath+r"\国家电网公司变电运维检修管理办法（试行）.xlsx")

    modifyfilecontent(rulepath+r"\国家电网公司变电检测管理规定（试行）.xlsx")
    modifyfilecontent(rulepath+r"\国家电网公司变电检修管理规定（试行）.xlsx")
    modifyfilecontent(rulepath+r"\国家电网公司变电评价管理规定（试行）.xlsx")
    modifyfilecontent(rulepath+r"\国家电网公司变电验收管理规定（试行）.xlsx")
    modifyfilecontent(rulepath+r"\国家电网公司变电运维管理规定（试行）.xlsx")
    modifyfilecontent(rulepath+r"\国家电网公司变电运维检修管理办法（试行）.xlsx")


